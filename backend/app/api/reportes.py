from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
from datetime import datetime, date
from typing import Optional

from app.core.database import get_db
from app.services.facturacion import (
    construir_reporte_facturacion
)
from app.services.presion import serializar_medicion, obtener_mediciones

from app.services.continuidad import construir_reporte_continuidad

from app.services.reclamos import construir_reporte_reclamos

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet





router = APIRouter(prefix="/reportes", tags=["Reportes"])



@router.get("/facturacion/{periodo}")
def reporte_facturacion(periodo: str, db: Session = Depends(get_db)):
    """
    Reporte de facturación con respaldo, por periodo, para fiscalización
    de la Superintendencia de Servicios Sanitarios (SISS).
    """
    try:
        return construir_reporte_facturacion(periodo, db)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.get("/facturacion/{periodo}/excel")
def reporte_facturacion_excel(periodo: str, db: Session = Depends(get_db)):
    try:
        reporte = construir_reporte_facturacion(periodo, db)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    wb = Workbook()
    ws = wb.active
    ws.title = f"Facturacion {periodo}"

    ws["A1"] = "Reporte de Facturación con Respaldo"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Periodo: {reporte['periodo']}"
    ws["A3"] = f"Tarifa vigente: {reporte['tarifa_vigente']}"
    ws["A4"] = f"Clientes facturados: {reporte['cantidad_clientes_facturados']}"
    ws["A5"] = f"Total recaudado: ${reporte['total_recaudado']:,.0f}"
    ws["A6"] = f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}"

    headers = [
        "RUT", "Nombre", "Dirección", "N° Medidor", "Socio", "Subsidio",
        "Fecha Lectura", "Lectura Anterior", "Lectura Actual", "Consumo m3",
        "Tarifa", "Cargo Fijo", "Monto Variable", "Subsidio Aplicado",
        "Subtotal Neto", "IVA Aplicado", "Total a Pagar"
    ]
    header_row = 8
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row = header_row + 1
    for r in reporte["detalle"]:
        ws.cell(row=row, column=1, value=r["rut"])
        ws.cell(row=row, column=2, value=r["nombre_cliente"])
        ws.cell(row=row, column=3, value=r["direccion"])
        ws.cell(row=row, column=4, value=r["numero_medidor"])
        ws.cell(row=row, column=5, value="Sí" if r["es_socio"] else "No")
        ws.cell(row=row, column=6, value="Sí" if r["tiene_subsidio"] else "No")
        ws.cell(row=row, column=7, value=str(r["fecha_lectura"]))
        ws.cell(row=row, column=8, value=r["lectura_anterior"])
        ws.cell(row=row, column=9, value=r["lectura_actual"])
        ws.cell(row=row, column=10, value=r["consumo_m3"])
        ws.cell(row=row, column=11, value=r["tarifa_aplicada"])
        ws.cell(row=row, column=12, value=r.get("cargo_fijo"))
        ws.cell(row=row, column=13, value=r.get("monto_variable"))
        ws.cell(row=row, column=14, value=r.get("subsidio_aplicado"))
        ws.cell(row=row, column=15, value=r.get("subtotal_neto"))
        ws.cell(row=row, column=16, value=r.get("iva_aplicado"))
        ws.cell(row=row, column=17, value=r["total_a_pagar"])
        row += 1

    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(length + 3, 30)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"reporte_facturacion_{periodo}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/facturacion/{periodo}/pdf")
def reporte_facturacion_pdf(periodo: str, db: Session = Depends(get_db)):
    try:
        reporte = construir_reporte_facturacion(periodo, db)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    styles = getSampleStyleSheet()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    elementos = []

    elementos.append(Paragraph("Reporte de Facturación con Respaldo", styles["Title"]))
    elementos.append(Paragraph(f"Periodo: {reporte['periodo']} | Tarifa vigente: {reporte['tarifa_vigente']}", styles["Normal"]))
    elementos.append(Paragraph(
        f"Clientes facturados: {reporte['cantidad_clientes_facturados']} | "
        f"Total recaudado: ${reporte['total_recaudado']:,.0f}", styles["Normal"]
    ))
    elementos.append(Paragraph(f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]))
    elementos.append(Spacer(1, 0.5 * cm))

    data = [["RUT", "Nombre", "N° Medidor", "Lect. Ant.", "Lect. Act.", "Consumo m3", "Tarifa", "Total a Pagar"]]
    for r in reporte["detalle"]:
        data.append([
            r["rut"], r["nombre_cliente"], r["numero_medidor"],
            r["lectura_anterior"], r["lectura_actual"], r["consumo_m3"],
            r["tarifa_aplicada"], f"${r['total_a_pagar']:,.0f}",
        ])

    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
    ]))
    elementos.append(tabla)

    doc.build(elementos)
    buffer.seek(0)

    filename = f"reporte_facturacion_{periodo}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )




@router.get("/presion/excel")
def reporte_presion_excel(
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
    db: Session = Depends(get_db),
):
    mediciones = obtener_mediciones(desde, hasta, db)
    datos = [serializar_medicion(m) for m in mediciones]

    wb = Workbook()
    ws = wb.active
    ws.title = "Registro de Presión"

    ws["A1"] = "Reporte de Presión de Servicio"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Desde: {desde or 'sin límite'}  Hasta: {hasta or 'sin límite'}"
    ws["A3"] = f"Mediciones: {len(datos)}"
    ws["A4"] = f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}"

    headers = ["Punto", "Ubicación", "Fecha", "Hora", "Presión (mca)", "Rango mínimo", "Rango máximo", "Cumple", "Observaciones"]
    header_row = 6
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row = header_row + 1
    for m in datos:
        ws.cell(row=row, column=1, value=m["punto_medicion"])
        ws.cell(row=row, column=2, value=m["ubicacion"])
        ws.cell(row=row, column=3, value=str(m["fecha_medicion"]))
        ws.cell(row=row, column=4, value=str(m["hora_medicion"]) if m["hora_medicion"] else None)
        ws.cell(row=row, column=5, value=m["presion_mca"])
        ws.cell(row=row, column=6, value=m["rango_minimo"])
        ws.cell(row=row, column=7, value=m["rango_maximo"])
        ws.cell(row=row, column=8, value="Sí" if m["cumple"] else "No")
        ws.cell(row=row, column=9, value=m["observaciones"])
        row += 1

    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(length + 3, 30)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_presion.xlsx"},
    )


@router.get("/presion/pdf")
def reporte_presion_pdf(
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
    db: Session = Depends(get_db),
):
    mediciones = obtener_mediciones(desde, hasta, db)
    datos = [serializar_medicion(m) for m in mediciones]
    styles = getSampleStyleSheet()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    elementos = []

    elementos.append(Paragraph("Reporte de Presión de Servicio", styles["Title"]))
    elementos.append(Paragraph(f"Desde: {desde or 'sin límite'} | Hasta: {hasta or 'sin límite'}", styles["Normal"]))
    elementos.append(Paragraph(f"Mediciones: {len(datos)}", styles["Normal"]))
    elementos.append(Paragraph(f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]))
    elementos.append(Spacer(1, 0.5 * cm))

    data = [["Punto", "Ubicación", "Fecha", "Presión (mca)", "Rango", "Cumple"]]
    for m in datos:
        data.append([
            m["punto_medicion"], m["ubicacion"] or "—", str(m["fecha_medicion"]),
            m["presion_mca"], f"{m['rango_minimo']}-{m['rango_maximo']}",
            "Sí" if m["cumple"] else "No",
        ])

    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
    ]))
    elementos.append(tabla)

    doc.build(elementos)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=reporte_presion.pdf"},
    )

@router.get("/continuidad/{periodo}")
def reporte_continuidad(periodo: str, db: Session = Depends(get_db)):
    """
    Reporte de continuidad de servicio (cortes y reposición), por periodo,
    para fiscalización de la Superintendencia de Servicios Sanitarios (SISS).
    """
    try:
        return construir_reporte_continuidad(periodo, db)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.get("/continuidad/{periodo}/excel")
def reporte_continuidad_excel(periodo: str, db: Session = Depends(get_db)):
    try:
        reporte = construir_reporte_continuidad(periodo, db)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    wb = Workbook()
    ws = wb.active
    ws.title = f"Continuidad {periodo}"

    ws["A1"] = "Reporte de Continuidad de Servicio"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Periodo: {reporte['periodo']}"
    ws["A3"] = f"Total cortes: {reporte['total_cortes']} (Activos: {reporte['total_activos']} / Cerrados: {reporte['total_cerrados']})"
    ws["A4"] = f"Duración total: {reporte['duracion_total_horas']} hrs | Duración promedio: {reporte['duracion_promedio_horas']} hrs"
    ws["A5"] = f"Clientes afectados (total): {reporte['total_clientes_afectados']}"
    ws["A6"] = f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}"

    headers = [
        "Estado", "Inicio", "Término", "Duración (hrs)", "Tipo", "Causa",
        "Sector Afectado", "Clientes Afectados", "Observaciones"
    ]
    header_row = 8
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row = header_row + 1
    todos = [("Activo", c) for c in reporte["cortes_activos"]] + \
            [("Cerrado", c) for c in reporte["cortes_cerrados"]]

    for estado, c in todos:
        ws.cell(row=row, column=1, value=estado)
        ws.cell(row=row, column=2, value=str(c["fecha_hora_inicio"]))
        ws.cell(row=row, column=3, value=str(c["fecha_hora_termino"]) if c["fecha_hora_termino"] else "—")
        ws.cell(row=row, column=4, value=c["duracion_horas"] if c["duracion_horas"] is not None else "—")
        ws.cell(row=row, column=5, value=c["tipo"])
        ws.cell(row=row, column=6, value=c["causa"])
        ws.cell(row=row, column=7, value=c["sector_afectado"])
        ws.cell(row=row, column=8, value=c["clientes_afectados"])
        ws.cell(row=row, column=9, value=c["observaciones"])
        row += 1

    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(length + 3, 30)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"reporte_continuidad_{periodo}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/continuidad/{periodo}/pdf")
def reporte_continuidad_pdf(periodo: str, db: Session = Depends(get_db)):
    try:
        reporte = construir_reporte_continuidad(periodo, db)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    styles = getSampleStyleSheet()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    elementos = []

    elementos.append(Paragraph("Reporte de Continuidad de Servicio", styles["Title"]))
    elementos.append(Paragraph(f"Periodo: {reporte['periodo']}", styles["Normal"]))
    elementos.append(Paragraph(
        f"Total cortes: {reporte['total_cortes']} (Activos: {reporte['total_activos']} / "
        f"Cerrados: {reporte['total_cerrados']})", styles["Normal"]
    ))
    elementos.append(Paragraph(
        f"Duración total: {reporte['duracion_total_horas']} hrs | "
        f"Duración promedio: {reporte['duracion_promedio_horas']} hrs | "
        f"Clientes afectados: {reporte['total_clientes_afectados']}", styles["Normal"]
    ))
    elementos.append(Paragraph(f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]))
    elementos.append(Spacer(1, 0.5 * cm))

    data = [["Estado", "Inicio", "Término", "Dur. (hrs)", "Tipo", "Sector", "Clientes Afect."]]
    todos = [("Activo", c) for c in reporte["cortes_activos"]] + \
            [("Cerrado", c) for c in reporte["cortes_cerrados"]]

    for estado, c in todos:
        data.append([
            estado,
            str(c["fecha_hora_inicio"]),
            str(c["fecha_hora_termino"]) if c["fecha_hora_termino"] else "—",
            c["duracion_horas"] if c["duracion_horas"] is not None else "—",
            c["tipo"],
            c["sector_afectado"],
            c["clientes_afectados"],
        ])

    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
    ]))
    elementos.append(tabla)

    doc.build(elementos)
    buffer.seek(0)

    filename = f"reporte_continuidad_{periodo}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )



@router.get("/reclamos/{periodo}")
def reporte_reclamos_json(periodo: str, db: Session = Depends(get_db)):
    return construir_reporte_reclamos(db, periodo)


@router.get("/reclamos/{periodo}/excel")
def reporte_reclamos_excel(periodo: str, db: Session = Depends(get_db)):
    reporte = construir_reporte_reclamos(db, periodo)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Reclamos {periodo}"

    ws["A1"] = "Libro de Reclamos"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Periodo: {reporte['periodo']}"
    ws["A3"] = f"Total reclamos: {reporte['total_reclamos']} (Respondidos: {reporte['total_respondidos']})"
    ws["A4"] = f"Fuera de plazo: {reporte['total_fuera_de_plazo']}"
    ws["A5"] = f"Promedio días hábiles de respuesta: {reporte['promedio_dias_habiles_respuesta'] if reporte['promedio_dias_habiles_respuesta'] is not None else '—'}"
    ws["A6"] = f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}"

    headers = [
        "Folio", "Tipo", "Reclamante", "Fecha Recepción", "Plazo Vencimiento",
        "Estado", "Fecha Respuesta", "Días Hábiles Respuesta", "Fuera de Plazo"
    ]
    header_row = 8
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row = header_row + 1
    for r in reporte["detalle"]:
        ws.cell(row=row, column=1, value=r["folio"])
        ws.cell(row=row, column=2, value=r["tipo_reclamo"])
        ws.cell(row=row, column=3, value=r["nombre_reclamante"] or "—")
        ws.cell(row=row, column=4, value=r["fecha_recepcion"])
        ws.cell(row=row, column=5, value=r["plazo_vencimiento"])
        ws.cell(row=row, column=6, value=r["estado"])
        ws.cell(row=row, column=7, value=r["fecha_respuesta"] or "—")
        ws.cell(row=row, column=8, value=r["dias_habiles_respuesta"] if r["dias_habiles_respuesta"] is not None else "—")
        ws.cell(row=row, column=9, value="Sí" if r["fuera_de_plazo"] else ("No" if r["fuera_de_plazo"] is not None else "—"))
        row += 1

    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(length + 3, 30)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"reporte_reclamos_{periodo}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/reclamos/{periodo}/pdf")
def reporte_reclamos_pdf(periodo: str, db: Session = Depends(get_db)):
    reporte = construir_reporte_reclamos(db, periodo)

    styles = getSampleStyleSheet()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    elementos = []

    elementos.append(Paragraph("Libro de Reclamos", styles["Title"]))
    elementos.append(Paragraph(f"Periodo: {reporte['periodo']}", styles["Normal"]))
    elementos.append(Paragraph(
        f"Total reclamos: {reporte['total_reclamos']} (Respondidos: {reporte['total_respondidos']}) | "
        f"Fuera de plazo: {reporte['total_fuera_de_plazo']}", styles["Normal"]
    ))
    promedio = reporte["promedio_dias_habiles_respuesta"]
    elementos.append(Paragraph(
        f"Promedio días hábiles de respuesta: {promedio if promedio is not None else '—'}", styles["Normal"]
    ))
    elementos.append(Paragraph(f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]))
    elementos.append(Spacer(1, 0.5 * cm))

    data = [["Folio", "Tipo", "Reclamante", "F. Recepción", "Plazo", "Estado", "Días Hábiles", "Fuera Plazo"]]
    for r in reporte["detalle"]:
        data.append([
            r["folio"],
            r["tipo_reclamo"],
            r["nombre_reclamante"] or "—",
            r["fecha_recepcion"],
            r["plazo_vencimiento"],
            r["estado"],
            r["dias_habiles_respuesta"] if r["dias_habiles_respuesta"] is not None else "—",
            "Sí" if r["fuera_de_plazo"] else ("No" if r["fuera_de_plazo"] is not None else "—"),
        ])

    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("ALIGN", (6, 1), (-1, -1), "CENTER"),
    ]))
    elementos.append(tabla)

    doc.build(elementos)
    buffer.seek(0)

    filename = f"reporte_reclamos_{periodo}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )