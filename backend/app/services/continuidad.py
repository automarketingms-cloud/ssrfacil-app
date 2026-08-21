from sqlalchemy.orm import Session
from sqlalchemy import func
from app.models.continuidad import CorteContinuidad
from app.schemas.continuidad import CorteCreate, CorteCierre
from datetime import date, datetime

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


def crear_corte(db: Session, corte: CorteCreate) -> CorteContinuidad:
    nuevo_corte = CorteContinuidad(**corte.model_dump())
    db.add(nuevo_corte)
    db.commit()
    db.refresh(nuevo_corte)
    return nuevo_corte


def cerrar_corte(db: Session, corte_id: int, cierre: CorteCierre) -> CorteContinuidad:
    corte = db.query(CorteContinuidad).filter(CorteContinuidad.id == corte_id).first()
    if not corte:
        return None
    if corte.fecha_hora_termino is not None:
        raise ValueError("Este corte ya fue cerrado")
    corte.fecha_hora_termino = cierre.fecha_hora_termino
    db.commit()
    db.refresh(corte)
    return corte


def calcular_duracion_horas(corte: CorteContinuidad) -> float | None:
    if corte.fecha_hora_termino is None:
        return None
    delta = corte.fecha_hora_termino - corte.fecha_hora_inicio
    return round(delta.total_seconds() / 3600, 2)


def listar_cortes(db: Session, periodo: str = None, solo_abiertos: bool = False):
    query = db.query(CorteContinuidad)
    if solo_abiertos:
        query = query.filter(CorteContinuidad.fecha_hora_termino.is_(None))
    if periodo:
        # periodo tipo "2026-07" -> rango [inicio_mes, inicio_mes_siguiente)
        anio, mes = map(int, periodo.split("-"))
        inicio = date(anio, mes, 1)
        fin = date(anio + 1, 1, 1) if mes == 12 else date(anio, mes + 1, 1)
        query = query.filter(
            CorteContinuidad.fecha_hora_inicio >= inicio,
            CorteContinuidad.fecha_hora_inicio < fin,
        )
    return query.order_by(CorteContinuidad.fecha_hora_inicio.desc()).all()

def contar_cortes_activos(db: Session) -> int:
    """Cuenta cortes sin reposición aún, sin traer los objetos completos."""
    return (
        db.query(func.count(CorteContinuidad.id))
        .filter(CorteContinuidad.fecha_hora_termino.is_(None))
        .scalar()
        or 0
    )


def serializar_corte(corte: CorteContinuidad) -> dict:
    return {
        "id": corte.id,
        "fecha_hora_inicio": corte.fecha_hora_inicio,
        "fecha_hora_termino": corte.fecha_hora_termino,
        "tipo": corte.tipo,
        "causa": corte.causa,
        "sector_afectado": corte.sector_afectado,
        "clientes_afectados": corte.clientes_afectados,
        "observaciones": corte.observaciones,
        "duracion_horas": calcular_duracion_horas(corte),
    }

def construir_reporte_continuidad(periodo: str, db: Session) -> dict:
    """
    Reporte de continuidad de servicio (cortes y reposición) para fiscalización SISS.
    Incluye cortes activos (aún sin reposición) y cortes cerrados del periodo,
    con el detalle de cuánto demoró cada reposición.
    """
    cortes = listar_cortes(db, periodo=periodo)

    if not cortes:
        raise ValueError(f"No hay cortes registrados para el periodo {periodo}")

    detalle_activos = []
    detalle_cerrados = []
    cortes_por_tipo: dict[str, int] = {}
    total_clientes_afectados = 0
    duraciones = []

    for c in cortes:
        data = serializar_corte(c)
        if c.fecha_hora_termino is None:
            detalle_activos.append(data)
        else:
            detalle_cerrados.append(data)
            if data["duracion_horas"] is not None:
                duraciones.append(data["duracion_horas"])

        cortes_por_tipo[c.tipo] = cortes_por_tipo.get(c.tipo, 0) + 1
        total_clientes_afectados += c.clientes_afectados or 0

    duracion_promedio = round(sum(duraciones) / len(duraciones), 2) if duraciones else 0
    duracion_total = round(sum(duraciones), 2) if duraciones else 0

    return {
        "periodo": periodo,
        "total_cortes": len(cortes),
        "total_activos": len(detalle_activos),
        "total_cerrados": len(detalle_cerrados),
        "duracion_promedio_horas": duracion_promedio,
        "duracion_total_horas": duracion_total,
        "total_clientes_afectados": total_clientes_afectados,
        "cortes_por_tipo": cortes_por_tipo,
        "cortes_activos": detalle_activos,
        "cortes_cerrados": detalle_cerrados,
    }

def construir_excel_reporte_continuidad(periodo: str, db: Session) -> BytesIO:
    reporte = construir_reporte_continuidad(periodo, db)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Continuidad {periodo}"

    ws["A1"] = "Reporte de Continuidad de Servicio"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Periodo: {reporte['periodo']}"
    ws["A3"] = f"Total cortes: {reporte['total_cortes']} (Activos: {reporte['total_activos']} / Cerrados: {reporte['total_cerrados']})"
    ws["A4"] = f"Duración total: {reporte['duracion_total_horas']} hrs | Duración promedio: {reporte['duracion_promedio_horas']} hrs"
    ws["A5"] = f"Clientes afectados (total): {reporte['total_clientes_afectados']}"
    ws["A6"] = f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}"

    headers = [
        "Estado", "Inicio", "Término", "Duración (hrs)", "Tipo", "Causa",
        "Sector Afectado", "Clientes Afectados", "Observaciones"
    ]
    header_row = 8
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row = header_row + 1
    todos = [("Activo", c) for c in reporte["cortes_activos"]] + \
            [("Cerrado", c) for c in reporte["cortes_cerrados"]]

    for estado, c in todos:
        ws.cell(row=row, column=1, value=estado)
        ws.cell(row=row, column=2, value=str(c["fecha_hora_inicio"]))
        ws.cell(row=row, column=3, value=str(c["fecha_hora_termino"]) if c["fecha_hora_termino"] else "—")
        ws.cell(row=row, column=4, value=c["duracion_horas"] if c["duracion_horas"] is not None else "—")
        ws.cell(row=row, column=5, value=c["tipo"])
        ws.cell(row=row, column=6, value=c["causa"])
        ws.cell(row=row, column=7, value=c["sector_afectado"])
        ws.cell(row=row, column=8, value=c["clientes_afectados"])
        ws.cell(row=row, column=9, value=c["observaciones"])
        row += 1

    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(length + 3, 30)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def construir_pdf_reporte_continuidad(periodo: str, db: Session) -> BytesIO:
    reporte = construir_reporte_continuidad(periodo, db)
    styles = getSampleStyleSheet()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    elementos = []

    elementos.append(Paragraph("Reporte de Continuidad de Servicio", styles["Title"]))
    elementos.append(Paragraph(f"Periodo: {reporte['periodo']}", styles["Normal"]))
    elementos.append(Paragraph(
        f"Total cortes: {reporte['total_cortes']} (Activos: {reporte['total_activos']} / "
        f"Cerrados: {reporte['total_cerrados']})", styles["Normal"]
    ))
    elementos.append(Paragraph(
        f"Duración total: {reporte['duracion_total_horas']} hrs | "
        f"Duración promedio: {reporte['duracion_promedio_horas']} hrs | "
        f"Clientes afectados: {reporte['total_clientes_afectados']}", styles["Normal"]
    ))
    elementos.append(Paragraph(f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]))
    elementos.append(Spacer(1, 0.5 * cm))

    data = [["Estado", "Inicio", "Término", "Dur. (hrs)", "Tipo", "Sector", "Clientes Afect."]]
    todos = [("Activo", c) for c in reporte["cortes_activos"]] + \
            [("Cerrado", c) for c in reporte["cortes_cerrados"]]

    for estado, c in todos:
        data.append([
            estado,
            str(c["fecha_hora_inicio"]),
            str(c["fecha_hora_termino"]) if c["fecha_hora_termino"] else "—",
            c["duracion_horas"] if c["duracion_horas"] is not None else "—",
            c["tipo"],
            c["sector_afectado"],
            c["clientes_afectados"],
        ])

    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
    ]))
    elementos.append(tabla)

    doc.build(elementos)
    buffer.seek(0)
    return buffer