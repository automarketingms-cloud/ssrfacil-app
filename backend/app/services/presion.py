from datetime import date, datetime
from typing import Optional
from io import BytesIO

from sqlalchemy.orm import Session

from app.models.presion import MedicionPresion

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

# Rango vigente: NCh 691:2015, DS MOP N°7 y N°14 de 2016 (15 a 70 mca)
# Rango anterior: según criterio interno indicado (8 a 40 mca antes de 2020)
# Si tienes el decreto exacto de este rango anterior, ajusta la fecha de corte.
FECHA_CAMBIO_NORMA = date(2020, 1, 1)
RANGO_ANTERIOR = (8, 40)
RANGO_VIGENTE = (15, 70)


def obtener_rango_normativo(fecha_medicion: date) -> tuple[float, float]:
    if fecha_medicion < FECHA_CAMBIO_NORMA:
        return RANGO_ANTERIOR
    return RANGO_VIGENTE


def evaluar_cumplimiento(presion_mca: float, fecha_medicion: date) -> dict:
    minimo, maximo = obtener_rango_normativo(fecha_medicion)
    return {
        "rango_minimo": minimo,
        "rango_maximo": maximo,
        "cumple": minimo <= presion_mca <= maximo,
    }


def serializar_medicion(m: MedicionPresion) -> dict:
    evaluacion = evaluar_cumplimiento(float(m.presion_mca), m.fecha_medicion)
    return {
        "id": m.id,
        "punto_medicion": m.punto_medicion,
        "ubicacion": m.ubicacion,
        "fecha_medicion": m.fecha_medicion,
        "hora_medicion": m.hora_medicion,
        "presion_mca": float(m.presion_mca),
        "observaciones": m.observaciones,
        "reclamo_id": m.reclamo_id,
        **evaluacion,
    }

def obtener_mediciones(desde: Optional[date], hasta: Optional[date], db: Session):
    query = db.query(MedicionPresion)
    if desde:
        query = query.filter(MedicionPresion.fecha_medicion >= desde)
    if hasta:
        query = query.filter(MedicionPresion.fecha_medicion <= hasta)
    return query.order_by(MedicionPresion.fecha_medicion.desc()).all()

def construir_excel_reporte_presion(desde: Optional[date], hasta: Optional[date], db: Session) -> BytesIO:
    mediciones = obtener_mediciones(desde, hasta, db)
    datos = [serializar_medicion(m) for m in mediciones]

    wb = Workbook()
    ws = wb.active
    ws.title = "Registro de Presión"

    ws["A1"] = "Reporte de Presión de Servicio"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Desde: {desde or 'sin límite'}  Hasta: {hasta or 'sin límite'}"
    ws["A3"] = f"Mediciones: {len(datos)}"
    ws["A4"] = f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}"

    headers = ["Punto", "Ubicación", "Fecha", "Hora", "Presión (mca)", "Rango mínimo", "Rango máximo", "Cumple", "Observaciones"]
    header_row = 6
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row = header_row + 1
    for m in datos:
        ws.cell(row=row, column=1, value=m["punto_medicion"])
        ws.cell(row=row, column=2, value=m["ubicacion"])
        ws.cell(row=row, column=3, value=str(m["fecha_medicion"]))
        ws.cell(row=row, column=4, value=str(m["hora_medicion"]) if m["hora_medicion"] else None)
        ws.cell(row=row, column=5, value=m["presion_mca"])
        ws.cell(row=row, column=6, value=m["rango_minimo"])
        ws.cell(row=row, column=7, value=m["rango_maximo"])
        ws.cell(row=row, column=8, value="Sí" if m["cumple"] else "No")
        ws.cell(row=row, column=9, value=m["observaciones"])
        row += 1

    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(length + 3, 30)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def construir_pdf_reporte_presion(desde: Optional[date], hasta: Optional[date], db: Session) -> BytesIO:
    mediciones = obtener_mediciones(desde, hasta, db)
    datos = [serializar_medicion(m) for m in mediciones]
    styles = getSampleStyleSheet()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    elementos = []

    elementos.append(Paragraph("Reporte de Presión de Servicio", styles["Title"]))
    elementos.append(Paragraph(f"Desde: {desde or 'sin límite'} | Hasta: {hasta or 'sin límite'}", styles["Normal"]))
    elementos.append(Paragraph(f"Mediciones: {len(datos)}", styles["Normal"]))
    elementos.append(Paragraph(f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]))
    elementos.append(Spacer(1, 0.5 * cm))

    data = [["Punto", "Ubicación", "Fecha", "Presión (mca)", "Rango", "Cumple"]]
    for m in datos:
        data.append([
            m["punto_medicion"], m["ubicacion"] or "—", str(m["fecha_medicion"]),
            m["presion_mca"], f"{m['rango_minimo']}-{m['rango_maximo']}",
            "Sí" if m["cumple"] else "No",
        ])

    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
    ]))
    elementos.append(tabla)

    doc.build(elementos)
    buffer.seek(0)
    return buffer