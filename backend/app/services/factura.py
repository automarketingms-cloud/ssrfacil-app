from datetime import date, timedelta, datetime
from sqlalchemy.orm import Session

from app.models.factura import Factura
from app.models.lectura import Lectura
from app.models.cliente import Cliente
from app.services.calculo_tarifa import (
    obtener_tarifa_vigente,
    calcular_consumo,
    calcular_total_a_pagar,
    validar_orden_periodo_facturacion,
)
from app.models.pago import Pago

from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from app.services.configuracion import obtener_configuracion
from app.services.pago import calcular_saldo_factura


def generar_factura(db: Session, cliente_id: int, periodo: str) -> Factura:
    ya_existe = (
        db.query(Factura)
        .filter(Factura.cliente_id == cliente_id, Factura.periodo == periodo)
        .first()
    )
    if ya_existe:
        raise ValueError("Ya existe una factura para este cliente en este periodo")

    validar_orden_periodo_facturacion(db, cliente_id, periodo)

    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise ValueError("Cliente no encontrado")

    lectura = (
        db.query(Lectura)
        .filter(Lectura.cliente_id == cliente_id, Lectura.periodo == periodo)
        .first()
    )
    if not lectura:
        raise ValueError("No hay lectura registrada para este cliente en este periodo")

    lectura_anterior_obj = (
        db.query(Lectura)
        .filter(Lectura.cliente_id == cliente_id, Lectura.periodo < periodo)
        .order_by(Lectura.periodo.desc())
        .first()
    )
    lectura_anterior_valor = lectura_anterior_obj.lectura_actual if lectura_anterior_obj else 0.0
    fecha_lectura_anterior = lectura_anterior_obj.fecha_lectura if lectura_anterior_obj else None

    tarifa = obtener_tarifa_vigente(db, periodo)

    viene_de_termino_medio = bool(lectura_anterior_obj and lectura_anterior_obj.es_promedio)
    consumo_medido = calcular_consumo(
        lectura.lectura_actual, lectura_anterior_valor, permitir_negativo=viene_de_termino_medio
    )

    consumo_a_facturar, mensaje_ajuste = aplicar_ajuste_credito_m3(cliente, consumo_medido)

    config = obtener_configuracion(db)

    desglose = calcular_total_a_pagar(consumo_a_facturar, tarifa, cliente, config.tasa_iva)

    fecha_emision = date.today()
    fecha_vencimiento = fecha_emision + timedelta(days=config.dias_plazo_pago)

    saldo_anterior, interes_mora = calcular_saldo_anterior_e_interes(
        db, cliente_id, periodo, config.tasa_interes_mora
    )
    mensaje_boleta_corte, fecha_limite_corte = calcular_aviso_corte(
        db, cliente_id, periodo, fecha_emision
    )
    mensaje_boleta = mensaje_ajuste or mensaje_boleta_corte

    factura = Factura(
        cliente_id=cliente_id,
        periodo=periodo,
         tipo_facturacion="termino_medio" if lectura.es_promedio else "normal",
        lectura_anterior=lectura_anterior_valor,
        lectura_actual=lectura.lectura_actual,
        fecha_lectura_anterior=fecha_lectura_anterior,
        fecha_lectura_actual=lectura.fecha_lectura,
        consumo_m3=consumo_a_facturar,
        detalle_tramos=desglose["detalle_tramos"],
        valor_fondo_reposicion=tarifa.valor_fondo_reposicion,
        cargo_fijo=desglose["cargo_fijo"],
        monto_variable=desglose["monto_variable"],
        cargo_fondo_reposicion=desglose["cargo_fondo_reposicion"],
        subsidio_aplicado=desglose["subsidio_aplicado"],
        porcentaje_subsidio_aplicado=(
            cliente.porcentaje_subsidio if desglose["subsidio_aplicado"] > 0 else None
        ),
        iva=desglose["iva_aplicado"],
        saldo_anterior=saldo_anterior,
        interes_mora=interes_mora,
        mensaje_boleta=mensaje_boleta,
        fecha_limite_corte=fecha_limite_corte,
        total_a_pagar=desglose["total_a_pagar"] + saldo_anterior + interes_mora,
        fecha_emision=fecha_emision,
        fecha_vencimiento=fecha_vencimiento,
        estado="pendiente",
    )
    db.add(cliente)  # persiste el credito_m3 actualizado
    db.add(factura)
    db.commit()
    db.refresh(factura)
    return factura

def calcular_saldo_anterior_e_interes(
    db: Session, cliente_id: int, periodo_actual: str, tasa_interes_anual: float
) -> tuple[float, float]:
    """
    Recorre las facturas anteriores impagas del cliente (cualquier periodo
    distinto al actual, con estado != 'pagada') y calcula:
      - saldo_anterior: la suma de TODO lo pendiente (venza o no), para que
        el cliente vea el total que arrastra.
      - interes_mora: interés calculado solo sobre la porción de ese saldo
        que ya está vencida (fecha_vencimiento < hoy), usando la tasa anual
        vigente en Configuración, prorrateada por los días de atraso.
    No modifica las facturas anteriores: es un cálculo de snapshot para la
    factura que se está generando ahora.
    """
    facturas_pendientes = (
        db.query(Factura)
        .filter(Factura.cliente_id == cliente_id, Factura.periodo != periodo_actual)
        .filter(Factura.estado != "pagada")
        .all()
    )

    saldo_anterior = 0.0
    interes_mora = 0.0
    hoy = date.today()

    for factura in facturas_pendientes:
        saldo = calcular_saldo_factura(db, factura)
        if saldo <= 0:
            continue

        saldo_anterior += saldo

        if hoy > factura.fecha_vencimiento:
            dias_atraso = (hoy - factura.fecha_vencimiento).days
            interes = saldo * (tasa_interes_anual / 100) / 365 * dias_atraso
            interes_mora += interes

    return round(saldo_anterior, 2), round(interes_mora, 2)

def calcular_aviso_corte(
    db: Session, cliente_id: int, periodo_actual: str, fecha_emision: date
) -> tuple[str | None, date | None]:
    """
    Determina si corresponde incluir un aviso de corte en la boleta que se
    está emitiendo: aplica si el cliente tiene alguna factura de un periodo
    distinto al actual, con estado != 'pagada' y saldo pendiente > 0, cuya
    fecha_vencimiento ya pasó a la fecha de emisión (es decir, está en mora).
    Plazo de aviso previo: 30 días (Ley 20.998).
    Devuelve (mensaje_boleta, fecha_limite_corte), o (None, None) si el
    cliente no está en mora.
    """
    facturas_pendientes = (
        db.query(Factura)
        .filter(Factura.cliente_id == cliente_id, Factura.periodo != periodo_actual)
        .filter(Factura.estado != "pagada")
        .all()
    )

    en_mora = any(
        factura.fecha_vencimiento < fecha_emision
        and calcular_saldo_factura(db, factura) > 0
        for factura in facturas_pendientes
    )

    if not en_mora:
        return None, None

    fecha_limite_corte = fecha_emision + timedelta(days=30)
    mensaje_boleta = (
        f"AVISO: Corte de servicio por no pago a partir del "
        f"{fecha_limite_corte.strftime('%d-%m-%Y')} si no regulariza su deuda."
    )
    return mensaje_boleta, fecha_limite_corte

def aplicar_ajuste_credito_m3(cliente: Cliente, consumo_medido: float) -> tuple[float, str | None]:
    """
    Resuelve el crédito de m3 acumulado por sobreestimaciones de término
    medio. Si consumo_medido es negativo (la lectura real, al volver, dio
    menos que el valor estimado), no se cobra ese mes y el excedente se
    guarda como crédito a favor del cliente. Si el cliente ya tenía
    crédito acumulado de antes, se descuenta primero del consumo de este
    mes (hasta agotar el crédito o el consumo, lo que sea menor), antes
    de calcular lo que se factura. Modifica cliente.credito_m3 in-place;
    el llamador es responsable de hacer commit.
    Devuelve (consumo_a_facturar, mensaje_para_boleta_o_None).
    """
    mensaje = None

    if consumo_medido < 0:
        credito_nuevo = -consumo_medido
        cliente.credito_m3 = round(cliente.credito_m3 + credito_nuevo, 2)
        mensaje = (
            f"Este período no se generó cobro por consumo: se generó un crédito de "
            f"{credito_nuevo:.0f} m3 a su favor (ajuste de un período estimado anterior), "
            f"que se descontará automáticamente de sus próximos consumos."
        )
        return 0.0, mensaje

    if cliente.credito_m3 > 0:
        credito_usado = min(cliente.credito_m3, consumo_medido)
        consumo_a_facturar = round(consumo_medido - credito_usado, 2)
        cliente.credito_m3 = round(cliente.credito_m3 - credito_usado, 2)
        mensaje = (
            f"Se descontaron {credito_usado:.0f} m3 de su crédito acumulado por ajuste de "
            f"período estimado anterior. Crédito restante: {cliente.credito_m3:.0f} m3."
        )
        return consumo_a_facturar, mensaje

    return consumo_medido, mensaje


def generar_facturas_periodo(db: Session, periodo: str) -> dict:
    """
    Genera facturas para todos los clientes con lectura registrada en el
    periodo que aún no tengan factura emitida. Bloquea el lote completo
    si aún no llega el día de facturación del mes. Si falla un cliente
    puntual (ej. sin tarifa vigente), no detiene el resto del lote: se
    registra en 'fallidas' y se sigue con los demás.
    """
    config = obtener_configuracion(db)

    periodo_actual = date.today().strftime("%Y-%m")
    es_periodo_actual = periodo == periodo_actual

    if es_periodo_actual and date.today().day < config.dia_facturacion:
        raise ValueError(
            f"Aún no se puede facturar el período actual: la emisión habilita desde el día {config.dia_facturacion} del mes"
        )

    lecturas = db.query(Lectura).filter(Lectura.periodo == periodo).all()
    generadas = []
    fallidas = []

    for lectura in lecturas:
        ya_existe = (
            db.query(Factura)
            .filter(
                Factura.cliente_id == lectura.cliente_id,
                Factura.periodo == periodo,
            )
            .first()
        )
        if ya_existe:
            continue

        try:
            factura = generar_factura(db, lectura.cliente_id, periodo)
            generadas.append(factura)
        except ValueError as e:
            fallidas.append({"cliente_id": lectura.cliente_id, "motivo": str(e)})

    return {
        "periodo": periodo,
        "generadas": generadas,
        "cantidad_generadas": len(generadas),
        "fallidas": fallidas,
        "cantidad_fallidas": len(fallidas),
    }


def obtener_factura(db: Session, factura_id: int) -> Factura:
    """
    Busca una factura por id. Lanza ValueError si no existe.
    """
    factura = db.query(Factura).filter(Factura.id == factura_id).first()
    if not factura:
        raise ValueError("Factura no encontrada")
    return factura


def listar_facturas(
    db: Session,
    periodo: str | None = None,
    cliente_id: int | None = None,
    estado: str | None = None,
) -> list[Factura]:
    """
    Lista facturas con filtros opcionales por periodo, cliente y estado.
    """
    query = db.query(Factura)
    if periodo:
        query = query.filter(Factura.periodo == periodo)
    if cliente_id:
        query = query.filter(Factura.cliente_id == cliente_id)
    if estado:
        query = query.filter(Factura.estado == estado)
    return query.order_by(Factura.fecha_emision.desc()).all()


def actualizar_facturas_vencidas(db: Session) -> int:
    """
    Marca como 'vencida' toda factura 'pendiente' cuya fecha_vencimiento
    ya pasó. Pensado para correrse periódicamente (ej. tarea diaria) o
    al entrar a la pantalla de facturación/morosos. Devuelve la cantidad
    de facturas actualizadas.
    """
    hoy = date.today()
    vencidas = (
        db.query(Factura)
        .filter(Factura.estado == "pendiente", Factura.fecha_vencimiento < hoy)
        .all()
    )
    for factura in vencidas:
        factura.estado = "vencida"
    db.commit()
    return len(vencidas)


def serializar_factura(factura: Factura, db: Session) -> dict:
    cliente = db.query(Cliente).filter(Cliente.id == factura.cliente_id).first()
    return {
        "id": factura.id,
        "cliente_id": factura.cliente_id,
        "nombre_cliente": cliente.nombre if cliente else None,
        "periodo": factura.periodo,
        "tipo_facturacion": factura.tipo_facturacion,
        "lectura_anterior": factura.lectura_anterior,
        "lectura_actual": factura.lectura_actual,
        "consumo_m3": factura.consumo_m3,
        "detalle_tramos": factura.detalle_tramos,
        "fecha_lectura_anterior": factura.fecha_lectura_anterior,
        "fecha_lectura_actual": factura.fecha_lectura_actual,
        "valor_fondo_reposicion": factura.valor_fondo_reposicion,
        "cargo_fijo": factura.cargo_fijo,
        "monto_variable": factura.monto_variable,
        "cargo_fondo_reposicion": factura.cargo_fondo_reposicion,
        "subsidio_aplicado": factura.subsidio_aplicado,
        "iva": factura.iva,
        "saldo_anterior": factura.saldo_anterior,
        "interes_mora": factura.interes_mora,
        "total_a_pagar": factura.total_a_pagar,
        "fecha_emision": factura.fecha_emision,
        "fecha_vencimiento": factura.fecha_vencimiento,
        "mensaje_boleta": factura.mensaje_boleta,
        "fecha_limite_corte": factura.fecha_limite_corte,
        "estado": factura.estado,
        "folio_sii": factura.folio_sii,
        "tipo_dte": factura.tipo_dte,
        "estado_envio_sii": factura.estado_envio_sii,
        "url_pdf_sii": factura.url_pdf_sii,
        "creado_en": factura.creado_en,
    }


def construir_pdf_factura(factura_id: int, db: Session) -> BytesIO:
    """
    Genera el PDF interno (respaldo propio, sin depender de SimpleAPI/SII)
    de una factura individual, usando el snapshot guardado al emitirla.
    """
    factura = obtener_factura(db, factura_id)
    cliente = db.query(Cliente).filter(Cliente.id == factura.cliente_id).first()
    config = obtener_configuracion(db)

    styles = getSampleStyleSheet()
    style_info = ParagraphStyle(
        "info", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#B00000")
    )
    style_small = ParagraphStyle("small", parent=styles["Normal"], fontSize=9, leading=13)
    style_small_bold = ParagraphStyle(
        "small_bold", parent=style_small, fontName="Helvetica-Bold"
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.2 * cm, bottomMargin=1.2 * cm)
    elementos = []

    # --- Encabezado: datos empresa (izq) + RUT/N° boleta (der) ---
    datos_empresa = [
        Paragraph(f"<b>{config.nombre_empresa}</b>", styles["Heading3"]),
        Paragraph(config.giro or "", style_small),
        Paragraph(config.direccion or "", style_small),
        Paragraph(f"Fono: {config.telefono or '—'}", style_small),
        Paragraph(f"Correo: {config.email or '—'}", style_small),
    ]
    datos_boleta = [
        Paragraph(f"RUT {config.rut_empresa}", style_small_bold),
        Paragraph("BOLETA INTERNA", style_small_bold),
        Paragraph(f"N° {factura.id}", style_small_bold),
    ]

    tabla_encabezado = Table(
        [[datos_empresa, datos_boleta]], colWidths=[11 * cm, 6 * cm]
    )
    tabla_encabezado.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOX", (1, 0), (1, 0), 1, colors.black),
        ("ALIGN", (1, 0), (1, 0), "CENTER"),
        ("TOPPADDING", (1, 0), (1, 0), 8),
        ("BOTTOMPADDING", (1, 0), (1, 0), 8),
    ]))
    elementos.append(tabla_encabezado)
    elementos.append(Spacer(1, 0.4 * cm))

    elementos.append(Paragraph(f"Periodo: {factura.periodo}", styles["Normal"]))
    elementos.append(Paragraph(f"Fecha emisión: {factura.fecha_emision.strftime('%d-%m-%Y')}", styles["Normal"]))
    elementos.append(Spacer(1, 0.4 * cm))

    # --- Datos usuario/socio ---
    datos_cliente = [
        [Paragraph("<b>DATOS USUARIO/SOCIO</b>", styles["Heading4"]), ""],
        [Paragraph(f"RUT: {cliente.rut}", style_small), Paragraph(f"N° Medidor: {cliente.numero_medidor}", style_small)],
        [Paragraph(f"Nombre: {cliente.nombre}", style_small), ""],
        [Paragraph(f"Dirección: {cliente.direccion}", style_small), ""],
    ]
    tabla_cliente = Table(datos_cliente, colWidths=[11 * cm, 6 * cm])
    tabla_cliente.setStyle(TableStyle([
        ("SPAN", (0, 0), (1, 0)),
        ("BOX", (0, 0), (-1, -1), 1, colors.black),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    elementos.append(tabla_cliente)
    elementos.append(Spacer(1, 0.5 * cm))

    # --- Columna izquierda: detalle en m3 ---
    fecha_lect_actual_str = (
        factura.fecha_lectura_actual.strftime("%d-%m-%y") if factura.fecha_lectura_actual else "—"
    )
    fecha_lect_anterior_str = (
        factura.fecha_lectura_anterior.strftime("%d-%m-%y") if factura.fecha_lectura_anterior else "—"
    )

    tramos_data = [["Tramo", "M3", "Valor ($)", "Total"]]
    for t in (factura.detalle_tramos or []):
        tramos_data.append([
            str(t["numero_tramo"]),
            f'{t["m3_en_tramo"]:.0f}',
            f'${t["precio_m3"]:,.0f}',
            f'${t["subtotal"]:,.0f}',
        ])

    total_m3_facturado = sum(t["m3_en_tramo"] for t in (factura.detalle_tramos or []))
    tramos_data.append(["Facturado", f"{total_m3_facturado:.0f}", "", f"${factura.monto_variable:,.0f}"])

    tabla_tramos = Table(tramos_data, colWidths=[2 * cm, 1.6 * cm, 2.2 * cm, 2.2 * cm])
    tabla_tramos.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.black),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.black),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F2F2F2")]),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
    ]))

    titulo_m3 = Table(
        [[Paragraph("<font color='white'><b>DETALLE DE CONSUMO EN M3</b></font>", styles["Normal"])]],
        colWidths=[9 * cm],
    )
    titulo_m3.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.black),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))

    col_izquierda = [
        titulo_m3,
        Spacer(1, 0.3 * cm),
        Paragraph(f"Lectura actual: {fecha_lect_actual_str}   {factura.lectura_actual:.0f}", style_small),
        Paragraph(f"Lectura anterior: {fecha_lect_anterior_str}   {factura.lectura_anterior:.0f}", style_small),
        Paragraph(f"<b>Consumo calculado: {factura.consumo_m3:.0f}</b>", style_small_bold),
        Spacer(1, 0.3 * cm),
        tabla_tramos,
        Spacer(1, 0.4 * cm),
        Paragraph(
            f"Aporte fondo de reposición y reinversión ($/m3): ${factura.valor_fondo_reposicion:,.0f}",
            style_small,
        ),
    ]

    # --- Columna derecha: detalle en pesos ---
    titulo_pesos = Table(
        [[Paragraph("<font color='white'><b>DETALLE DE CONSUMO EN PESOS</b></font>", styles["Normal"])]],
        colWidths=[8 * cm],
    )
    titulo_pesos.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.black),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))

    pesos_data = [
        ["Cargo fijo", f"${factura.cargo_fijo:,.0f}"],
        ["Consumo (Cargo variable AP)", f"${factura.monto_variable:,.0f}"],
        ["Cargo fondo de reposición", f"${factura.cargo_fondo_reposicion:,.0f}"],
    ]
    if factura.subsidio_aplicado > 0:
        pesos_data.append(["Subsidio aplicado", f"-${factura.subsidio_aplicado:,.0f}"])
    if factura.iva > 0:
        pesos_data.append(["IVA", f"${factura.iva:,.0f}"])

    subtotal = (
        factura.cargo_fijo + factura.monto_variable + factura.cargo_fondo_reposicion
        - factura.subsidio_aplicado + factura.iva
    )
    pesos_data.append(["SUBTOTAL", f"${subtotal:,.0f}"])

    tabla_pesos = Table(pesos_data, colWidths=[5.5 * cm, 2.5 * cm])
    tabla_pesos.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LINEBELOW", (0, -2), (-1, -2), 0.5, colors.grey),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    col_derecha = [titulo_pesos, Spacer(1, 0.3 * cm), tabla_pesos]

    if factura.saldo_anterior > 0 or factura.interes_mora > 0:
        col_derecha.append(Spacer(1, 0.4 * cm))
        arrastre_data = []
        if factura.saldo_anterior > 0:
            arrastre_data.append(["Saldo anterior", f"${factura.saldo_anterior:,.0f}"])
        if factura.interes_mora > 0:
            arrastre_data.append(["Interés por mora", f"${factura.interes_mora:,.0f}"])
        tabla_arrastre = Table(arrastre_data, colWidths=[5.5 * cm, 2.5 * cm])
        tabla_arrastre.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ]))
        col_derecha.append(tabla_arrastre)

    tabla_columnas = Table([[col_izquierda, col_derecha]], colWidths=[9.2 * cm, 8 * cm])
    tabla_columnas.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elementos.append(tabla_columnas)
    elementos.append(Spacer(1, 0.6 * cm))

    # --- Total a pagar destacado ---
    total_data = [
        ["TOTAL A PAGAR", f"${factura.total_a_pagar:,.0f}"],
        ["VENCIMIENTO", factura.fecha_vencimiento.strftime("%d-%m-%Y")],
    ]
    tabla_total = Table(total_data, colWidths=[9.2 * cm, 4 * cm])
    tabla_total.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#E5E5E5")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 12),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elementos.append(tabla_total)

    # --- Estado / recuadro de información (mensaje boleta, aviso de corte) ---
    elementos.append(Spacer(1, 0.5 * cm))
    elementos.append(Paragraph(f"Estado: {factura.estado.upper()}", styles["Normal"]))

    if factura.mensaje_boleta:
        elementos.append(Spacer(1, 0.3 * cm))
        info_data = [[Paragraph("<b>Información</b>", styles["Normal"])],
                     [Paragraph(factura.mensaje_boleta, style_info)]]
        if factura.fecha_limite_corte:
            info_data.append([Paragraph(
                f"Fecha límite antes de posible corte: {factura.fecha_limite_corte.strftime('%d-%m-%Y')}",
                style_info
            )])
        tabla_info = Table(info_data, colWidths=[13 * cm])
        tabla_info.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#B00000")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ]))
        elementos.append(tabla_info)

    doc.build(elementos)
    buffer.seek(0)
    return buffer


def construir_reporte_facturacion(periodo: str, db: Session) -> dict:
    """
    Arma el reporte de facturación con respaldo para un periodo dado.
    Usado por el endpoint JSON y los exports (Excel/PDF).
    Lee directo de la tabla Factura (fuente de verdad, snapshot congelado
    al emitir), así el reporte siempre coincide exactamente con la boleta
    real que recibió cada cliente. Requiere que las facturas del periodo
    ya hayan sido generadas; lanza ValueError si no hay ninguna.

    Optimizado para evitar consultas N+1: en vez de una query a Cliente,
    Factura (anterior) y Pago por cada factura del período, se cargan
    todas de una vez (3 queries grandes) y se cruzan en memoria.
    """
    facturas = db.query(Factura).filter(Factura.periodo == periodo).all()
    if not facturas:
        raise ValueError(
            "No hay facturas emitidas para este periodo. "
            "Genera las facturas en el módulo de Facturación primero."
        )

    tarifa = obtener_tarifa_vigente(db, periodo)
    config = obtener_configuracion(db)

    cliente_ids = [f.cliente_id for f in facturas]

    # --- Carga masiva de clientes ---
    clientes_por_id = {
        c.id: c
        for c in db.query(Cliente).filter(Cliente.id.in_(cliente_ids)).all()
    }

    # --- Carga masiva de la última factura anterior de cada cliente ---
    # (para saber desde qué fecha contar los pagos "del período")
    facturas_previas = (
        db.query(Factura)
        .filter(Factura.cliente_id.in_(cliente_ids), Factura.periodo < periodo)
        .order_by(Factura.cliente_id, Factura.periodo.desc())
        .all()
    )
    factura_anterior_por_cliente: dict[int, Factura] = {}
    for f in facturas_previas:
        # como viene ordenado por periodo desc, la primera que aparece
        # por cliente_id es la más reciente
        factura_anterior_por_cliente.setdefault(f.cliente_id, f)

    # --- Carga masiva de todos los pagos relevantes ---
    pagos_rows = (
        db.query(Pago, Factura.cliente_id)
        .join(Factura, Pago.factura_id == Factura.id)
        .filter(Factura.cliente_id.in_(cliente_ids))
        .all()
    )
    pagos_por_cliente: dict[int, list[Pago]] = {}
    for pago, cid in pagos_rows:
        pagos_por_cliente.setdefault(cid, []).append(pago)

    detalle = []
    total_recaudado = 0.0

    for factura in facturas:
        cliente = clientes_por_id.get(factura.cliente_id)
        if not cliente:
            continue

        subtotal = round(
            factura.cargo_fijo + factura.monto_variable + factura.cargo_fondo_reposicion, 2
        )
        corte_en_tramite = bool(
            factura.fecha_limite_corte
            and factura.estado != "pagada"
            and date.today() >= factura.fecha_limite_corte
        )

        factura_anterior = factura_anterior_por_cliente.get(factura.cliente_id)
        fecha_desde_pagos = factura_anterior.fecha_emision if factura_anterior else date.min

        pagos_cliente = pagos_por_cliente.get(factura.cliente_id, [])
        pagos_del_periodo = [
            p for p in pagos_cliente
            if fecha_desde_pagos < p.fecha_pago <= factura.fecha_emision
        ]
        monto_pagado_periodo = round(sum(p.monto for p in pagos_del_periodo), 2)
        fecha_ultimo_pago = max((p.fecha_pago for p in pagos_del_periodo), default=None)

        registro = {
            "cliente_id": cliente.id,
            "nombre_cliente": cliente.nombre,
            "rut": cliente.rut,
            "direccion": cliente.direccion,
            "numero_medidor": cliente.numero_medidor,
            "es_socio": cliente.es_socio,
            "tiene_subsidio": cliente.tiene_subsidio,
            "periodo": periodo,
            "tipo_facturacion": factura.tipo_facturacion,
            "fecha_lectura_actual": factura.fecha_lectura_actual,
            "fecha_lectura_anterior": factura.fecha_lectura_anterior,
            "lectura_anterior": factura.lectura_anterior,
            "lectura_actual": factura.lectura_actual,
            "consumo_m3": factura.consumo_m3,
            "tarifa_aplicada": tarifa.nombre,
            "cargo_fijo": factura.cargo_fijo,
            "monto_variable": factura.monto_variable,
            "cargo_fondo_reposicion": factura.cargo_fondo_reposicion,
            "subtotal": subtotal,
            "subsidio_aplicado": factura.subsidio_aplicado,
            "porcentaje_subsidio_aplicado": factura.porcentaje_subsidio_aplicado,
            "m3_subsidiados": (
                next((t["m3_en_tramo"] for t in (factura.detalle_tramos or []) if t["numero_tramo"] == 1), 0)
                if factura.subsidio_aplicado > 0 else 0
            ),
            "subtotal_neto": round(subtotal - factura.subsidio_aplicado, 2),
            "iva_aplicado": factura.iva,
            "saldo_anterior": factura.saldo_anterior,
            "interes_mora": factura.interes_mora,
            "total_a_pagar": factura.total_a_pagar,
            "fecha_emision": factura.fecha_emision,
            "fecha_vencimiento": factura.fecha_vencimiento,
            "corte_en_tramite": corte_en_tramite,
            "estado": factura.estado,
            "monto_pagado_periodo": monto_pagado_periodo,
            "fecha_ultimo_pago": fecha_ultimo_pago,
        }
        detalle.append(registro)
        total_recaudado += registro["total_a_pagar"]

    return {
        "periodo": periodo,
        "tarifa_vigente": tarifa.nombre,
        "cantidad_clientes_facturados": len(detalle),
        "total_recaudado": round(total_recaudado, 2),
        "telefono_atencion": config.telefono,
        "horario_atencion": config.horario_atencion,
        "detalle": detalle,
    }

def construir_excel_reporte_facturacion(periodo: str, db: Session) -> BytesIO:
    """
    Genera el Excel del reporte de facturación con respaldo (fiscalización SISS).
    """
    reporte = construir_reporte_facturacion(periodo, db)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Facturacion {periodo}"

    ws["A1"] = "Reporte de Facturación con Respaldo"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Periodo: {reporte['periodo']}"
    ws["A3"] = f"Tarifa vigente: {reporte['tarifa_vigente']}"
    ws["A4"] = f"Clientes facturados: {reporte['cantidad_clientes_facturados']}"
    ws["A5"] = f"Total recaudado: ${reporte['total_recaudado']:,.0f}"
    ws["A6"] = f"Teléfono de atención: {reporte['telefono_atencion'] or '—'}"
    ws["A7"] = f"Horario de atención: {reporte['horario_atencion'] or '—'}"
    ws["A8"] = f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}"

    headers = [
        "RUT", "Nombre", "Dirección", "N° Medidor", "Socio", "Subsidio",
        "Tipo Facturación", "Fecha Lect. Anterior", "Fecha Lect. Actual",
        "Lectura Anterior", "Lectura Actual", "Consumo m3",
        "Tarifa", "Cargo Fijo", "Monto Variable", "Subtotal",
        "M3 Subsidiados", "% Subsidio", "Subsidio Aplicado",
        "Subtotal Neto", "IVA Aplicado", "Saldo Anterior", "Interés Mora",
        "Total a Pagar", "Vencimiento", "Monto Pagado (periodo)",
        "Fecha Último Pago"
    ]
    header_row = 10
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row = header_row + 1
    for r in reporte["detalle"]:
        ws.cell(row=row, column=1, value=r["rut"])
        ws.cell(row=row, column=2, value=r["nombre_cliente"])
        ws.cell(row=row, column=3, value=r["direccion"])
        ws.cell(row=row, column=4, value=r["numero_medidor"])
        ws.cell(row=row, column=5, value="Sí" if r["es_socio"] else "No")
        ws.cell(row=row, column=6, value="Sí" if r["tiene_subsidio"] else "No")
        ws.cell(row=row, column=7, value=r["tipo_facturacion"])
        ws.cell(row=row, column=8, value=str(r["fecha_lectura_anterior"]) if r.get("fecha_lectura_anterior") else "—")
        ws.cell(row=row, column=9, value=str(r["fecha_lectura_actual"]) if r.get("fecha_lectura_actual") else "—")
        ws.cell(row=row, column=10, value=r["lectura_anterior"])
        ws.cell(row=row, column=11, value=r["lectura_actual"])
        ws.cell(row=row, column=12, value=r["consumo_m3"])
        ws.cell(row=row, column=13, value=r["tarifa_aplicada"])
        ws.cell(row=row, column=14, value=r.get("cargo_fijo"))
        ws.cell(row=row, column=15, value=r.get("monto_variable"))
        ws.cell(row=row, column=16, value=r.get("subtotal"))
        ws.cell(row=row, column=17, value=r.get("m3_subsidiados") or "—")
        ws.cell(row=row, column=18, value=r.get("porcentaje_subsidio_aplicado") or "—")
        ws.cell(row=row, column=19, value=f"-{r['subsidio_aplicado']:,.0f}" if r.get("subsidio_aplicado") else "—")
        ws.cell(row=row, column=20, value=r.get("subtotal_neto"))
        ws.cell(row=row, column=21, value=r.get("iva_aplicado"))
        ws.cell(row=row, column=22, value=r.get("saldo_anterior") or 0)
        ws.cell(row=row, column=23, value=r.get("interes_mora") or 0)
        ws.cell(row=row, column=24, value=r["total_a_pagar"])
        ws.cell(
            row=row, column=25,
            value="Corte en Trámite" if r["corte_en_tramite"]
            else (str(r["fecha_vencimiento"]) if r["fecha_vencimiento"] else "—")
        )
        ws.cell(row=row, column=26, value=r.get("monto_pagado_periodo") or 0)
        ws.cell(row=row, column=27, value=str(r["fecha_ultimo_pago"]) if r.get("fecha_ultimo_pago") else "—")
        row += 1

    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(length + 3, 30)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def construir_pdf_reporte_facturacion(periodo: str, db: Session) -> BytesIO:
    """
    Genera el PDF del reporte de facturación con respaldo (fiscalización SISS).
    """
    reporte = construir_reporte_facturacion(periodo, db)
    styles = getSampleStyleSheet()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    elementos = []

    elementos.append(Paragraph("Reporte de Facturación con Respaldo", styles["Title"]))
    elementos.append(Paragraph(f"Periodo: {reporte['periodo']} | Tarifa vigente: {reporte['tarifa_vigente']}", styles["Normal"]))
    elementos.append(Paragraph(
        f"Clientes facturados: {reporte['cantidad_clientes_facturados']} | "
        f"Total recaudado: ${reporte['total_recaudado']:,.0f}", styles["Normal"]
    ))
    elementos.append(Paragraph(
        f"Teléfono: {reporte['telefono_atencion'] or '—'} | Horario: {reporte['horario_atencion'] or '—'}",
        styles["Normal"]
    ))
    elementos.append(Paragraph(f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]))
    elementos.append(Spacer(1, 0.5 * cm))

    data = [["RUT", "Nombre", "N° Medidor", "Consumo m3", "Subtotal", "Subsidio", "Saldo Anterior", "Total a Pagar", "Vencimiento"]]
    for r in reporte["detalle"]:
        vencimiento = (
            "Corte en Trámite" if r["corte_en_tramite"]
            else (str(r["fecha_vencimiento"]) if r["fecha_vencimiento"] else "—")
        )
        data.append([
            r["rut"],
            r["nombre_cliente"],
            r["numero_medidor"],
            r["consumo_m3"],
            f"${r.get('subtotal', 0):,.0f}",
            f"-${r['subsidio_aplicado']:,.0f}" if r.get("subsidio_aplicado") else "—",
            f"${r['saldo_anterior']:,.0f}" if r.get("saldo_anterior") else "—",
            f"${r['total_a_pagar']:,.0f}",
            vencimiento,
        ])

    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
    ]))
    elementos.append(tabla)

    doc.build(elementos)
    buffer.seek(0)
    return buffer