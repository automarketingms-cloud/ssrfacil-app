from sqlalchemy.orm import Session
from sqlalchemy import extract
from datetime import datetime, date, timedelta
import holidays

from app.models.reclamo import Reclamo
from app.models.cliente import Cliente 
from app.schemas.reclamo import ReclamoCreate, ReclamoResponder

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

DIAS_HABILES_PLAZO = 10
feriados_chile = holidays.CL()


def es_dia_habil(fecha: date) -> bool:
    return fecha.weekday() < 5 and fecha not in feriados_chile  # 0=lunes ... 4=viernes


def sumar_dias_habiles(fecha_inicio: date, dias: int) -> date:
    """Suma 'dias' días hábiles a partir de fecha_inicio (sin contar fecha_inicio)."""
    fecha = fecha_inicio
    contados = 0
    while contados < dias:
        fecha += timedelta(days=1)
        if es_dia_habil(fecha):
            contados += 1
    return fecha


def contar_dias_habiles_entre(fecha_inicio: date, fecha_fin: date) -> int:
    """Cuenta los días hábiles entre dos fechas (sin contar fecha_inicio, incluyendo fecha_fin)."""
    if fecha_fin <= fecha_inicio:
        return 0
    dias = 0
    fecha = fecha_inicio
    while fecha < fecha_fin:
        fecha += timedelta(days=1)
        if es_dia_habil(fecha):
            dias += 1
    return dias


def generar_folio(db: Session, anio: int) -> str:
    """Genera el siguiente folio correlativo para el año, ej '2026-001'."""
    ultimo = (
        db.query(Reclamo)
        .filter(extract("year", Reclamo.fecha_recepcion) == anio)
        .order_by(Reclamo.id.desc())
        .first()
    )
    if ultimo is None:
        siguiente = 1
    else:
        siguiente = int(ultimo.folio.split("-")[1]) + 1
    return f"{anio}-{siguiente:03d}"


def crear_reclamo(db: Session, datos: ReclamoCreate) -> Reclamo:
    fecha_recepcion = datos.fecha_recepcion or datetime.now()
    anio = fecha_recepcion.year
    folio = generar_folio(db, anio)
    plazo_vencimiento = sumar_dias_habiles(fecha_recepcion.date(), DIAS_HABILES_PLAZO)

    nombre_reclamante = datos.nombre_reclamante
    rut_reclamante = datos.rut_reclamante

    if datos.cliente_id is not None:
        cliente = db.query(Cliente).filter(Cliente.id == datos.cliente_id).first()
        if cliente is None:
            raise ValueError("Cliente no encontrado")
        # snapshot: se copian aunque el reclamo no traiga estos campos explícitos
        nombre_reclamante = cliente.nombre
        rut_reclamante = cliente.rut

    reclamo = Reclamo(
        folio=folio,
        anio=anio,
        cliente_id=datos.cliente_id,
        nombre_reclamante=nombre_reclamante,
        rut_reclamante=rut_reclamante,
        direccion_reclamo=datos.direccion_reclamo,
        tipo_reclamo=datos.tipo_reclamo,
        descripcion=datos.descripcion,
        fecha_recepcion=fecha_recepcion,
        plazo_vencimiento=plazo_vencimiento,
        estado="abierto",
        observaciones=datos.observaciones,
    )
    db.add(reclamo)
    db.commit()
    db.refresh(reclamo)
    return reclamo


def responder_reclamo(db: Session, reclamo_id: int, datos: ReclamoResponder) -> Reclamo:
    reclamo = db.query(Reclamo).filter(Reclamo.id == reclamo_id).first()
    if reclamo is None:
        return None

    fecha_respuesta = datetime.now()
    dias_habiles_respuesta = contar_dias_habiles_entre(
        reclamo.fecha_recepcion.date(), fecha_respuesta.date()
    )
    fuera_de_plazo = fecha_respuesta.date() > reclamo.plazo_vencimiento

    reclamo.respuesta = datos.respuesta
    reclamo.fecha_respuesta = fecha_respuesta
    reclamo.dias_habiles_respuesta = dias_habiles_respuesta
    reclamo.fuera_de_plazo = fuera_de_plazo
    reclamo.estado = "respondido"

    db.commit()
    db.refresh(reclamo)
    return reclamo


def cerrar_reclamo(db: Session, reclamo_id: int) -> Reclamo:
    """Cierra un reclamo que ya fue respondido."""
    reclamo = db.query(Reclamo).filter(Reclamo.id == reclamo_id).first()
    if reclamo is None:
        return None
    if reclamo.estado != "respondido":
        raise ValueError("Solo se puede cerrar un reclamo que ya fue respondido")
    reclamo.estado = "cerrado"
    db.commit()
    db.refresh(reclamo)
    return reclamo


def cerrar_reclamo_sin_respuesta(db: Session, reclamo_id: int, motivo: str) -> Reclamo:
    """Cierra un reclamo directamente, sin pasar por 'respondido' (ej. retiro, duplicado)."""
    reclamo = db.query(Reclamo).filter(Reclamo.id == reclamo_id).first()
    if reclamo is None:
        return None
    if reclamo.estado != "abierto":
        raise ValueError("Solo se puede aplicar cierre directo a un reclamo abierto")
    reclamo.estado = "cerrado_sin_respuesta"
    reclamo.motivo_cierre = motivo
    db.commit()
    db.refresh(reclamo)
    return reclamo


def listar_reclamos(db: Session, periodo: str = None, estado: str = None, cliente_id: int = None):
    """periodo en formato 'YYYY-MM', filtra por fecha_recepcion."""
    query = db.query(Reclamo)
    if periodo:
        anio, mes = periodo.split("-")
        query = query.filter(
            extract("year", Reclamo.fecha_recepcion) == int(anio),
            extract("month", Reclamo.fecha_recepcion) == int(mes),
        )
    if estado:
        query = query.filter(Reclamo.estado == estado)
    if cliente_id:
        query = query.filter(Reclamo.cliente_id == cliente_id)
    return query.order_by(Reclamo.fecha_recepcion.desc()).all()

def obtener_reclamo(db: Session, reclamo_id: int) -> Reclamo:
    return db.query(Reclamo).filter(Reclamo.id == reclamo_id).first()


def construir_reporte_reclamos(db: Session, periodo: str) -> dict:
    """periodo en formato 'YYYY-MM'. Arma el resumen para el reporte de fiscalización SISS."""
    reclamos = listar_reclamos(db, periodo=periodo)

    total = len(reclamos)
    respondidos = [r for r in reclamos if r.dias_habiles_respuesta is not None]
    fuera_de_plazo = [r for r in respondidos if r.fuera_de_plazo]

    por_tipo: dict[str, int] = {}
    por_estado: dict[str, int] = {}
    for r in reclamos:
        por_tipo[r.tipo_reclamo] = por_tipo.get(r.tipo_reclamo, 0) + 1
        por_estado[r.estado] = por_estado.get(r.estado, 0) + 1

    promedio_dias_respuesta = (
        round(sum(r.dias_habiles_respuesta for r in respondidos) / len(respondidos), 1)
        if respondidos
        else None
    )

    detalle = [
        {
            "folio": r.folio,
            "tipo_reclamo": r.tipo_reclamo,
            "nombre_reclamante": r.nombre_reclamante,
            "fecha_recepcion": r.fecha_recepcion.strftime("%Y-%m-%d"),
            "plazo_vencimiento": r.plazo_vencimiento.strftime("%Y-%m-%d"),
            "estado": r.estado,
            "fecha_respuesta": r.fecha_respuesta.strftime("%Y-%m-%d") if r.fecha_respuesta else None,
            "dias_habiles_respuesta": r.dias_habiles_respuesta,
            "fuera_de_plazo": r.fuera_de_plazo,
        }
        for r in reclamos
    ]

    return {
        "periodo": periodo,
        "total_reclamos": total,
        "total_respondidos": len(respondidos),
        "total_fuera_de_plazo": len(fuera_de_plazo),
        "promedio_dias_habiles_respuesta": promedio_dias_respuesta,
        "reclamos_por_tipo": por_tipo,
        "reclamos_por_estado": por_estado,
        "detalle": detalle,
    }

def construir_excel_reporte_reclamos(periodo: str, db: Session) -> BytesIO:
    reporte = construir_reporte_reclamos(db, periodo)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Reclamos {periodo}"

    ws["A1"] = "Libro de Reclamos"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Periodo: {reporte['periodo']}"
    ws["A3"] = f"Total reclamos: {reporte['total_reclamos']} (Respondidos: {reporte['total_respondidos']})"
    ws["A4"] = f"Fuera de plazo: {reporte['total_fuera_de_plazo']}"
    ws["A5"] = f"Promedio días hábiles de respuesta: {reporte['promedio_dias_habiles_respuesta'] if reporte['promedio_dias_habiles_respuesta'] is not None else '—'}"
    ws["A6"] = f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}"

    headers = [
        "Folio", "Tipo", "Reclamante", "Fecha Recepción", "Plazo Vencimiento",
        "Estado", "Fecha Respuesta", "Días Hábiles Respuesta", "Fuera de Plazo"
    ]
    header_row = 8
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row = header_row + 1
    for r in reporte["detalle"]:
        ws.cell(row=row, column=1, value=r["folio"])
        ws.cell(row=row, column=2, value=r["tipo_reclamo"])
        ws.cell(row=row, column=3, value=r["nombre_reclamante"] or "—")
        ws.cell(row=row, column=4, value=r["fecha_recepcion"])
        ws.cell(row=row, column=5, value=r["plazo_vencimiento"])
        ws.cell(row=row, column=6, value=r["estado"])
        ws.cell(row=row, column=7, value=r["fecha_respuesta"] or "—")
        ws.cell(row=row, column=8, value=r["dias_habiles_respuesta"] if r["dias_habiles_respuesta"] is not None else "—")
        ws.cell(row=row, column=9, value="Sí" if r["fuera_de_plazo"] else ("No" if r["fuera_de_plazo"] is not None else "—"))
        row += 1

    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(length + 3, 30)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def construir_pdf_reporte_reclamos(periodo: str, db: Session) -> BytesIO:
    reporte = construir_reporte_reclamos(db, periodo)
    styles = getSampleStyleSheet()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    elementos = []

    elementos.append(Paragraph("Libro de Reclamos", styles["Title"]))
    elementos.append(Paragraph(f"Periodo: {reporte['periodo']}", styles["Normal"]))
    elementos.append(Paragraph(
        f"Total reclamos: {reporte['total_reclamos']} (Respondidos: {reporte['total_respondidos']}) | "
        f"Fuera de plazo: {reporte['total_fuera_de_plazo']}", styles["Normal"]
    ))
    promedio = reporte["promedio_dias_habiles_respuesta"]
    elementos.append(Paragraph(
        f"Promedio días hábiles de respuesta: {promedio if promedio is not None else '—'}", styles["Normal"]
    ))
    elementos.append(Paragraph(f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]))
    elementos.append(Spacer(1, 0.5 * cm))

    data = [["Folio", "Tipo", "Reclamante", "F. Recepción", "Plazo", "Estado", "Días Hábiles", "Fuera Plazo"]]
    for r in reporte["detalle"]:
        data.append([
            r["folio"],
            r["tipo_reclamo"],
            r["nombre_reclamante"] or "—",
            r["fecha_recepcion"],
            r["plazo_vencimiento"],
            r["estado"],
            r["dias_habiles_respuesta"] if r["dias_habiles_respuesta"] is not None else "—",
            "Sí" if r["fuera_de_plazo"] else ("No" if r["fuera_de_plazo"] is not None else "—"),
        ])

    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("ALIGN", (6, 1), (-1, -1), "CENTER"),
    ]))
    elementos.append(tabla)

    doc.build(elementos)
    buffer.seek(0)
    return buffer